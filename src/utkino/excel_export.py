from __future__ import annotations

import csv
from datetime import datetime, time
from pathlib import Path
from typing import Any, Iterable


def export_workbook_to_csv(input_file: Path, output_dir: Path) -> list[Path]:
    """Export every workbook sheet to CSV without Excel/COM automation."""
    suffix = input_file.suffix.lower()
    output_dir.mkdir(parents=True, exist_ok=True)

    if suffix == ".xls":
        return _export_xls(input_file, output_dir)
    if suffix in {".xlsx", ".xlsm"}:
        return _export_xlsx(input_file, output_dir)

    raise ValueError(
        f"Unsupported Excel format: {input_file.suffix}. Use .xls, .xlsx, or .xlsm."
    )


def _export_xls(input_file: Path, output_dir: Path) -> list[Path]:
    try:
        import xlrd
    except ImportError as error:
        raise RuntimeError("Install xlrd to read .xls files: python -m pip install xlrd") from error

    workbook = xlrd.open_workbook(str(input_file), formatting_info=False)
    written: list[Path] = []

    for sheet in workbook.sheets():
        path = output_dir / f"{sheet.name}.csv"
        rows = (
            [
                _format_xls_cell(sheet.cell(row_index, column_index), workbook.datemode)
                for column_index in range(sheet.ncols)
            ]
            for row_index in range(sheet.nrows)
        )
        _write_csv(path, rows)
        written.append(path)

    return written


def _export_xlsx(input_file: Path, output_dir: Path) -> list[Path]:
    try:
        import openpyxl
    except ImportError as error:
        raise RuntimeError(
            "Install openpyxl to read .xlsx files: python -m pip install openpyxl"
        ) from error

    workbook = openpyxl.load_workbook(
        input_file,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    written: list[Path] = []

    try:
        for sheet in workbook.worksheets:
            path = output_dir / f"{sheet.title}.csv"
            rows = (
                [_format_value(value) for value in row]
                for row in sheet.iter_rows(values_only=True)
            )
            _write_csv(path, rows)
            written.append(path)
    finally:
        workbook.close()

    return written


def _write_csv(path: Path, rows: Iterable[list[str]]) -> None:
    with path.open("w", encoding="cp1251", errors="replace", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(rows)


def _format_xls_cell(cell: Any, datemode: int) -> str:
    import xlrd

    if cell.ctype == xlrd.XL_CELL_EMPTY:
        return ""
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return _format_value(xlrd.xldate.xldate_as_datetime(cell.value, datemode))
        except (OverflowError, ValueError):
            return _format_value(cell.value)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return "TRUE" if cell.value else "FALSE"
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return ""
    return _format_value(cell.value)


def _format_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.time() == time.min:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value)
